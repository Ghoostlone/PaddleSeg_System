import os.path

import cv2
import pymysql  # MySQL包
from flask import Flask, request, session, redirect, send_file  # 后端Flask包
from flask import render_template
from flask_bootstrap import Bootstrap  # 前端Bootstrap包
from openpyxl.styles import Alignment  # 表格包
from werkzeug.utils import secure_filename  # 文件读取包
import shutil
from gevent import pywsgi  # gevent库中的服务器，用于运行Flask应用程序。
import vtk  # 用于创建、渲染和分析3D图像的库。
from trimesh.exchange.obj import export_obj  # 用于将3D模型导出为其他格式的函数。
from openpyxl import load_workbook  # 用于加载现有的Excel工作簿。
from openpyxl.utils import get_column_letter  # 用于将列索引转换为Excel中的字母列标识。


# 几个用于nii转ply的工具函数（read_nii、get_mc_contour、smoothing、singledisplay、multidisplay、write_ply）
def read_nii(filename):
    """
    读取nii文件，输入文件路径
    """
    # 创建一个vtkNIFTIImageReader对象
    reader = vtk.vtkNIFTIImageReader()
    # 设置要读取的NIfTI文件的文件名
    reader.SetFileName(filename)
    # 执行读取操作
    reader.Update()
    # 返回VTK图像读取器
    return reader


def get_mc_contour(file, setvalue):
    """
    计算轮廓的方法
    file:读取的vtk类
    setvalue:要得到的轮廓的值
    """
    contour = vtk.vtkDiscreteMarchingCubes()

    # ChatGPT 解释:
    # vtkDiscreteMarchingCubes和vtkMarchingCubes是两个在VTK（Visualization Toolkit）库中用于构建3D模型的算法。它们的主要区别在于如何处理输入数据。
    # vtkMarchingCubes是一种连续的算法，用于从体数据（例如体素数据或标量场数据）中提取等值面。
    # 它基于Marching Cubes算法，将体数据划分为小立方体单元，并根据单元内部的数值情况确定等值面的形状和拓扑关系。
    # vtkMarchingCubes可以处理连续的数据集，并产生光滑的等值面。
    # 与之相反，vtkDiscreteMarchingCubes是一种离散的算法，用于从离散的体数据中提取等值面。
    # 它适用于离散的数据集，其中体素只能具有预定义的几种状态，例如0和1。离散数据集通常用于表示二值图像或分割结果等。
    # vtkDiscreteMarchingCubes使用类似的原理，但对于离散数据集，它仅考虑预定义状态之间的界面，并生成离散的等值面。
    # 总结一下，vtkMarchingCubes适用于处理连续的体数据，可以产生光滑的等值面，
    # 而vtkDiscreteMarchingCubes适用于处理离散的体数据，只考虑预定义状态之间的界面，生成离散的等值面。
    # 选择使用哪个算法取决于您的数据类型和应用需求。

    # 设置输入数据连接到vtkDiscreteMarchingCubes
    contour.SetInputConnection(file.GetOutputPort())
    # 打开法线计算以生成法线信息
    contour.ComputeNormalsOn()
    # 设置要提取的等值面数值阈值
    contour.SetValue(0, setvalue)
    # 返回生成的轮廓对象
    return contour


def smoothing(smoothing_iterations, pass_band, feature_angle, contour):
    '''
    使轮廓变平滑
    smoothing_iterations:迭代次数
    pass_band:值越小单次平滑效果越明显
    feature_angle:特征角度参数，暂时不知道作用
    contour (vtk.vtkDiscreteMarchingCubes): 输入的轮廓对象
    '''
    # vtk有两种平滑函数，效果类似

    vtk.vtkSmoothPolyDataFilter()
    smoother = vtk.vtkSmoothPolyDataFilter()  # 创建平滑滤波器
    smoother.SetInputConnection(contour.GetOutputPort())  # 连接输入轮廓到平滑滤波器
    smoother.SetNumberOfIterations(50)  # 设置平滑迭代次数
    smoother.SetRelaxationFactor(0.6)  # 越大效果越明显

    vtk.vtkWindowedSincPolyDataFilter()
    smoother = vtk.vtkWindowedSincPolyDataFilter()
    smoother.SetInputConnection(contour.GetOutputPort())
    smoother.SetNumberOfIterations(smoothing_iterations)
    smoother.BoundarySmoothingOff()
    smoother.FeatureEdgeSmoothingOff()
    smoother.SetFeatureAngle(feature_angle)  # 暂时不清楚feature_angle参数的作用
    smoother.SetPassBand(pass_band)  # 设置控制平滑强度的参数，值越小单次平滑效果越明显
    smoother.NonManifoldSmoothingOn()
    smoother.NormalizeCoordinatesOn()
    smoother.Update()  # 执行平滑操作
    return smoother  # 返回平滑后的轮廓对象


def singledisplay(obj):
    # 创建一个VTK映射器，将输入对象连接到映射器
    mapper = vtk.vtkPolyDataMapper()
    mapper.SetInputConnection(obj.GetOutputPort())
    mapper.ScalarVisibilityOff()
    # 创建一个VTKActor，将映射器连接到actor
    actor = vtk.vtkActor()
    actor.SetMapper(mapper)
    # 创建一个VTK渲染器，设置背景颜色并添加actor
    renderer = vtk.vtkRenderer()
    renderer.SetBackground([0.1, 0.1, 0.5])
    renderer.AddActor(actor)
    # 创建一个VTK渲染窗口，设置窗口大小并添加渲染器
    window = vtk.vtkRenderWindow()
    window.SetSize(512, 512)
    window.AddRenderer(renderer)
    # 创建一个VTK渲染窗口交互器，用于与用户的交互
    interactor = vtk.vtkRenderWindowInteractor()
    interactor.SetRenderWindow(window)

    # 开始显示可视化
    window.Render()
    interactor.Initialize()
    interactor.Start()
    export_obj(window)
    return window


def multidisplay(obj):
    # 给每个零部件上色并显示
    # This sets the block at flat index 3 red
    # Note that the index is the flat index in the tree, so the whole multiblock
    # is index 0 and the blocks are flat indexes 1, 2 and 3.  This affects
    # the block returned by mbds.GetBlock(2).
    colors = vtk.vtkNamedColors()
    # 创建一个VTK CompositePolyDataMapper2 映射器，设置输入对象
    mapper = vtk.vtkCompositePolyDataMapper2()
    mapper.SetInputDataObject(obj)
    # 创建一个VTK CompositeDataDisplayAttributes 对象，用于指定各个块的显示属性
    cdsa = vtk.vtkCompositeDataDisplayAttributes()
    mapper.SetCompositeDataDisplayAttributes(cdsa)
    # 上色
    mapper.SetBlockColor(1, colors.GetColor3d('Red'))
    mapper.SetBlockColor(2, colors.GetColor3d('Lavender'))
    mapper.SetBlockColor(3, colors.GetColor3d('Gray'))
    mapper.SetBlockColor(4, colors.GetColor3d('Green'))
    mapper.SetBlockColor(5, colors.GetColor3d('Yellow'))
    mapper.SetBlockColor(6, colors.GetColor3d('Pink'))
    mapper.SetBlockColor(7, colors.GetColor3d('Brown'))
    mapper.SetBlockColor(8, colors.GetColor3d('Turquoise'))
    mapper.SetBlockColor(9, colors.GetColor3d('Orange'))
    mapper.SetBlockColor(10, colors.GetColor3d('Blue'))
    mapper.SetBlockColor(11, colors.GetColor3d('Purple'))
    actor = vtk.vtkActor()
    actor.SetMapper(mapper)

    # 创建渲染器（Renderer）、渲染窗口（RenderWindow）和渲染窗口交互器（RenderWindowInteractor）
    renderer = vtk.vtkRenderer()
    renderWindow = vtk.vtkRenderWindow()
    renderWindow.AddRenderer(renderer)
    renderWindowInteractor = vtk.vtkRenderWindowInteractor()
    renderWindowInteractor.SetRenderWindow(renderWindow)

    # 启用用户界面交互器（UI Interactor）
    renderer.AddActor(actor)
    renderer.SetBackground(colors.GetColor3d('SteelBlue'))
    renderWindow.SetWindowName('CompositePolyDataMapper')
    renderWindow.Render()
    renderWindowInteractor.Start()


def write_ply(obj, save_dir, color):
    """
    输入必须是单个模型，vtkMultiBlockDataSet没有GetOutputPort()类
    """

    plyWriter = vtk.vtkPLYWriter()  # 创建一个VTK PLY写入器
    plyWriter.SetFileName(save_dir)  # 设置保存路径和文件名
    plyWriter.SetColorModeToUniformCellColor()  # 设置颜色模式为统一单元颜色
    plyWriter.SetColor(color[0], color[1], color[2])  # 设置颜色
    plyWriter.SetInputConnection(obj.GetOutputPort())  # 将输入模型连接到PLY写入器
    plyWriter.Write()  # 执行PLY文件写入


# 初始化Flask后端
app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(5)
bootstrap = Bootstrap(app)

# 连接到数据库
cnn = pymysql.connect(host="", port=64563, user="root", password="", database="test", charset="utf8")
cursor = cnn.cursor()


# 登录页面
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template("login/login.html")
    if request.method == 'POST':
        inputId = request.form.get('inputId')
        inputPassword = request.form.get('inputPassword')
        print(inputId, inputPassword)
        cursor.execute("SELECT pwd , identity FROM `user` WHERE id='" + inputId + "'")  # 身份验证
        result = cursor.fetchall()
        print(result)
        if result:
            for row in result:
                if row[0] == inputPassword:
                    session['userid'] = inputId
                    session['identity'] = row[1]
                    return redirect('/index/index')
            else:
                return render_template("login/wrongPWD.html")
        else:
            return render_template("/login/loginFail.html")


# 注册页面
@app.route('/signup', methods=['POST', 'GET'])
def signup():
    if request.method == 'GET':
        return render_template("signup/signup.html")
        # return redirect("https://www.baidu.com")
    if request.method == 'POST':
        # 处理POST请求，接收用户提交的注册信息
        inputId = request.form.get("inputId")  # 获取用户输入的ID
        inputEmail = request.form.get("inputEmail")  # 获取用户输入的邮箱
        inputPassword = request.form.get("inputPassword")  # 获取用户输入的密码
        identity_Flag = request.form.get("identity")  # 获取用户选择的身份标识
        Identity = ""  # 初始化身份
        Sir_key = request.form.get("sir_yes_sir")  # 获取管理员密钥（如果是医生注册）

        # 根据用户选择的身份标识设置身份
        if identity_Flag == "0":
            Identity = "doctor"
        elif identity_Flag == "1":
            Identity = "patient"
        # 注册信息打印调试
        print("receive sign up request:" + inputId, inputEmail, inputPassword, Identity)
        # 查询重复用户
        cursor.execute("SELECT * FROM USER WHERE id = '%s'" % inputId)
        results = cursor.fetchall()
        if results:
            return render_template("signup/SignupFail.html")
        else:
            sql = "INSERT INTO  `user` VALUES('" + inputId + "','" + inputPassword + "','" \
                  + inputEmail + "','" + Identity + "')"
            if identity_Flag == "0":  # 如果用户选择身份为医生，需要管理员密钥验证
                if Sir_key == "LGDLGD":  # 此处为管理员密钥，记得设置为超参数或从数据库获取多级密钥
                    n = cursor.execute(sql)
                    cursor.connection.commit()
                    print(inputId + " signing up successfully")
                    return render_template("signup/SignupSuccess.html")
                else:
                    return render_template("signup/SignupFail.html")
            elif identity_Flag == "1":  # 如果用户选择身份为患者，直接插入用户信息
                n = cursor.execute(sql)
                cursor.connection.commit()
                print(inputId + " signing up successfully")
                return render_template("signup/SignupSuccess.html")


# 主页
@app.route('/index/index', methods=['GET'])
def index():
    if request.method == 'GET':
        if session.get('identity') == "doctor":
            return render_template("index/for_doctor.html", id=session.get('userid'))
        elif session.get('identity') == "patient":
            return render_template("index/for_patient.html", id=session.get('userid'))


# 上传CT
@app.route('/upload/upload', methods=['POST', 'GET'])
def upload():
    if request.method == 'GET':
        return render_template("upload/upload_image.html", id=session.get('userid'))
    if request.method == 'POST':
        f = request.files['image']  # 获取上传的文件
        diagnosis_ID = request.form.get("diagnosis")  # 获取诊断ID
        patient_ID = request.form.get("id")  # 获取病人ID
        base_path = os.path.dirname(__file__)  # 获取当前目录
        print(base_path)
        # 检查病人ID是否匹配
        cursor.execute("SELECT * FROM `user` WHERE id='" + patient_ID + "'")
        result = cursor.fetchall()
        print(result)
        if result:
            for row in result:
                if row[3] == "patient":
                    # 生成nnunet所需要的文件名(文件本名_0000.nii.gz)
                    img_path = secure_filename(f.filename).split('.', 1)[0] + "_0000.nii.gz"
                    print(img_path)
                    dir_path = os.path.join(base_path, 'static/img/nii_path', patient_ID)
                    print(dir_path)
                    upload_path = os.path.join(base_path, 'static/img/nii_path', patient_ID, img_path)
                    print(upload_path)
                    if os.path.exists(dir_path):
                        f.save(upload_path)
                    else:
                        os.mkdir(dir_path)
                        f.save(upload_path)
                    # 打印步骤信息方便确认
                    print(diagnosis_ID, session.get('userid'), patient_ID, upload_path)
                    print(upload_path)
                    upload_path = upload_path.replace("\\", "/")
                    sql = "INSERT INTO diagnosis VALUES (" + diagnosis_ID + ",'" + session.get(
                        'userid') + "'," + patient_ID + ",'" + upload_path + "')        "
                    cursor.execute(sql)
                    cursor.connection.commit()
                    print(123132132132132132121213123213213)
                    return "已上传成功"
                else:
                    return "此ID非病人，请重新输入"
        else:
            return "查无此人，请重新输入病人ID"


# 看CT
@app.route('/CT_view/ct_view', methods=['GET', 'POST'])
def CT_view():
    if request.method == 'GET':
        return render_template("CT_view/ct_view.html", id=session.get('userid'))
    if request.method == 'POST':
        P_ID = request.form.get("P_ID")  # 获取病人ID
        ply_file_path = "./static/img/ply_path/" + P_ID + "/"  # 此路径用于查询CT是否存在
        ply_path = "../static/img/ply_path/" + P_ID + "/"  # 此路径用于渲染CT部件到WEB
        print(ply_path)
        if os.path.exists(ply_file_path):
            return render_template("CT_view/3D_render.html", id=session.get('userid'), ply_path=ply_path)
        else:
            return render_template("CT_view/3D_ERROR.html")


@app.route('/start_Predict', methods=['GET', 'POST'])
def start_Predict():
    if request.method == 'GET':
        return render_template("Predict/start_Predict.html", id=session.get('userid'))
    if request.method == 'POST':  # 用于ajax的查询此病人所有CT文件路径的路由
        P_ID = request.form.get("id")
        print(P_ID)
        session['patient_id_selected'] = P_ID
        cursor.execute("SELECT * FROM `diagnosis` WHERE patient_id='" + P_ID + "'")
        result = cursor.fetchall()
        print(result)
        all_row = ""
        if result:
            for row in result:
                all_row = all_row + row[3] + "<br />"
            return all_row
        else:
            return 124141412412412412412412412412  # 如果没有相关的诊断记录，打印此特征码
        # command_Line = "python nnunet/infer.py --image_folder /root/autodl-tmp/Flask/static/img/nii_path/"
        # os.system(command_Line)


@app.route('/run_predict', methods=['POST'])
def run_Pred():
    if request.method == 'POST':
        File_path = request.form.get("file_path")
        # 设置文件路径
        session['File'] = File_path.split('/', )
        final_path = File_path.split('.', 1)[0] + ".nii.gz"
        final_path = final_path.replace('nii_path', 'seg_path')
        final_path = final_path.replace('_0000.nii.gz', '.nii.gz')
        session['final_path'] = final_path
        print(session.get('final_path'))
        return render_template("Predict/running.html")


@app.route('/PaddleSeg', methods=['GET'])
def PaddleSeg():
    if request.method == 'GET':
        P_ID_SELECTED = session.get('patient_id_selected')
        file_path = "/root/autodl-tmp/Flask/static/img/nii_path/" + P_ID_SELECTED + "/"
        output_path = "/root/autodl-tmp/Flask/static/img/seg_path/" + P_ID_SELECTED + "/"
        if os.path.exists(output_path):
            print("有了")
        else:
            os.mkdir(output_path)
            print("没有，创了")
        # 一切设置完毕，开始使用静态模型预测
        # python路径+训练参数
        SegCommand = "/root/miniconda3/envs/PaddleSeg/bin/python " \
                     "/root/autodl-tmp/Flask/PaddleSeg/contrib/MedicalSeg/nnunet/infer.py --image_folder " + file_path + " --output_folder " + output_path + " " \
                                                                                                                                                             "--plan_path /root/autodl-tmp/Flask/predict/nnUNetPlansv2.1_plans_3D.pkl " \
                                                                                                                                                             "--model_paths /root/autodl-tmp/Flask/predict/baseline_model/model.pdmodel " \
                                                                                                                                                             "--param_paths /root/autodl-tmp/Flask/predict/baseline_model/model.pdiparams " \
                                                                                                                                                             "--postprocessing_json_path /root/autodl-tmp/Flask/predict/baseline_model/postprocessing.json " \
                                                                                                                                                             "--model_type cascade_lowres " \
                                                                                                                                                             "--disable_postprocessing " \
                                                                                                                                                             "--save_npz"
        os.system("cd /root/autodl-tmp/Flask/PaddleSeg/contrib/MedicalSeg")
        os.chdir("/root/autodl-tmp/Flask/PaddleSeg/contrib/MedicalSeg")
        print(os.getcwd())
        os.system(SegCommand)
        os.chdir("/root/autodl-tmp/Flask")
        # 分割完毕，开始转nii为ply保存
        print("这是final_path", session.get("final_path"))
        nii_dir = session.get('final_path')  # 获取最终的NII文件路径
        ply_path = "./static/img/ply_path/" + P_ID_SELECTED + "/"  # PLY文件的保存路径
        if os.path.exists(ply_path):
            print("有了")
        else:
            os.mkdir(ply_path)
            print("没有，创了")
        save_dir = ply_path
        smoothing_iterations = 100  # 平滑迭代次数
        pass_band = 0.005  # 通带带宽
        feature_angle = 120  # 特征角度
        reader = read_nii(nii_dir)  # 读取NII文件

        color = [(0, 0, 0), (128, 0, 0), (0, 128, 0), (128, 128, 0), (0, 0, 128), (128, 0, 128), (0, 128, 128),
                 (128, 128, 128), (64, 0, 0), (192, 0, 0), (64, 128,
                                                            0), (192, 128, 0), (64, 0, 128), (192, 0, 128),
                 (64, 128, 128), (192, 128, 128), (0, 64, 0), (128, 64, 0), (0, 192, 0), (128, 192, 0), (0, 64, 128),
                 (128, 64, 12)]  # 颜色信息

        mbds = vtk.vtkMultiBlockDataSet()
        mbds.SetNumberOfBlocks(11)
        items = ['background', '1', '2',
                 '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
        for iter in range(1, 12):
            print(iter)
            contour = get_mc_contour(reader, iter)  # 获取轮廓
            smoothing_iterations = 100
            pass_band = 0.005
            feature_angle = 120
            smoother = smoothing(smoothing_iterations, pass_band,
                                 feature_angle, contour)  # 进行平滑处理
            write_ply(smoother, save_dir + f'{items[iter]}.ply', color[iter])  # 将PLY文件保存到指定路径

            mbds.SetBlock(iter, smoother.GetOutput())  # 将平滑后的结果添加到MultiBlockDataSet
        # 以下是用于调试处理（用ply浏览软件监控分割结果）
        # singledisplay(smoother)
        # write_ply(mbds, save_dir + f'final.ply', color[3])
        # multidisplay(mbds)
        return render_template("index/for_doctor.html", id=session.get('userid'))


@app.route('/P_CT', methods=['POST', 'GET'])  # 病人查看本人CT的页面
def p_ct():
    if request.method == 'GET':
        return render_template("CT_view/ct_view_P.html", id=session.get('userid'))
    if request.method == 'POST':
        P_ID = session.get('userid')
        base_path = os.path.dirname(__file__)
        print(base_path)
        ply_path = "./static/img/ply_path/" + P_ID + "/"
        upload_path = os.path.join(base_path, ply_path)
        print(upload_path)
        return render_template("CT_view/3D_render.html", id=session.get('userid'), ply_path=ply_path)


@app.route('/search', methods=['POST'])  # 用于ajax查询病人ID信息的路由，只接收POST请求
def search():
    if request.method == 'POST':
        cursor.execute("SELECT * FROM `user` WHERE identity='patient'")
        result = cursor.fetchall()
        return_things = ""
        if result:
            # print(result)
            for i in result:
                # print(i[0])
                return_things = return_things + "&nbsp&nbsp&nbsp" + i[0]
        return return_things


@app.route('/tomesh', methods=['GET'])  # 暂时用不到
def tomesh():
    if request.method == 'GET':
        print(12123123)
    if request.method == 'POST':
        cursor.execute("SELECT * FROM `user` WHERE identity='patient'")
        result = cursor.fetchall()
        return_things = ""
        if result:
            # print(result)
            for i in result:
                # print(i[0])
                return_things = return_things + "&nbsp&nbsp&nbsp" + i[0]
        return return_things


@app.route("/form", methods=['GET', 'POST'])  # 用于生成表格型诊断报告
def formtest():
    if request.method == 'GET':
        return render_template("form/form.html", id=session.get('userid'))
    if request.method == 'POST':
        # 加载模板文件
        template_file = 'static/xlsx/standard.xlsx'
        wb = load_workbook(template_file)

        # 选择要填充数据的工作表
        sheet = wb['Sheet1']

        # 准备要填充的数据（用每格的首字母一一定位）
        XM = request.form.get("XM")
        ID = request.form.get("ID")
        ZSZD = request.form.get("ZSZD")
        ZSYZ = request.form.get("ZSYZ")
        PZD = request.form.get("PZD")
        PYZ = request.form.get("PYZ")
        GZD = request.form.get("GZD")
        GYZ = request.form.get("GYZ")
        YXZD = request.form.get("YXZD")
        YXYZ = request.form.get("YXYZ")
        ZDMZD = request.form.get("ZDMZD")
        ZDMYZ = request.form.get("ZDMYZ")
        PGZD = request.form.get("PGZD")
        PGYZ = request.form.get("PGYZ")
        WZD = request.form.get("WZD")
        WYZ = request.form.get("WYZ")
        DNZD = request.form.get("DNZD")
        DNYZ = request.form.get("DNYZ")
        XQJMZD = request.form.get("XQJMZD")
        XQJMYZ = request.form.get("XQJMYZ")
        SGZD = request.form.get("SGZD")
        SGYZ = request.form.get("SGYZ")
        YSZD = request.form.get("YSZD")
        YSYZ = request.form.get("YSYZ")
        # 填充数据到工作表中
        sheet.cell(3, 5).value = XM
        sheet.cell(5, 3).value = ZSZD
        sheet.cell(5, 5).value = ZSYZ
        sheet.cell(6, 3).value = PZD
        sheet.cell(6, 5).value = PYZ
        sheet.cell(7, 3).value = GZD
        sheet.cell(7, 5).value = GYZ
        sheet.cell(8, 3).value = YXZD
        sheet.cell(8, 5).value = YXYZ
        sheet.cell(9, 3).value = ZDMZD
        sheet.cell(9, 5).value = ZDMYZ
        sheet.cell(10, 3).value = PGZD
        sheet.cell(10, 5).value = PGYZ
        sheet.cell(11, 3).value = WZD
        sheet.cell(11, 5).value = WYZ
        sheet.cell(12, 3).value = DNZD
        sheet.cell(12, 5).value = DNYZ
        sheet.cell(13, 3).value = XQJMZD
        sheet.cell(13, 5).value = XQJMYZ
        sheet.cell(14, 3).value = SGZD
        sheet.cell(14, 5).value = SGYZ
        sheet.cell(15, 3).value = YSZD
        sheet.cell(15, 5).value = YSYZ

        # 设置第3列和第5列的单元格文本自动换行
        for i in range(5, 15):
            sheet.cell(i, 3).alignment = Alignment(wrap_text=True)
            sheet.cell(i, 5).alignment = Alignment(wrap_text=True)
        # cell = sheet.cell(row=row, column=col)
        # 自动调整每列的宽度，使文本适应单元格宽度
        for col in range(1, 5):
            col_letter = get_column_letter(col)
            sheet.column_dimensions[col_letter].auto_size = True

        # 创建输出目录
        output_file = "static/xlsx/" + ID + "/"
        if os.path.exists(output_file):
            print("有了")
        else:
            os.mkdir(output_file)
            print("没有，创了")
        # 保存为新的Excel文件
        output_path = output_file + "output.xlsx"
        wb.save(output_path)

        print("数据已成功填充到Excel文件中。")
        return render_template("form/form.html", id=session.get('userid'))


@app.route('/form_download', methods=['POST', 'GET'])  # 下载诊断报告
def F_Download():
    if request.method == 'GET':
        return render_template("form/form_Download.html", id=session.get('userid'))
    if request.method == 'POST':
        ID = request.form.get("ID")
        path = "static/xlsx/" + ID + "/output.xlsx"
        if (os.path.exists(path)):
            print("文档搜到")
            return send_file(path, as_attachment=True)
        else:
            return render_template("form/form_ERROR.html")


@app.route('/form_download_P', methods=['POST', 'GET'])  # 病人下载诊断报告
def F_Download_P():
    if request.method == 'GET':
        ID = session.get("userid")
        path = "static/xlsx/" + ID + "/output.xlsx"
        print("文档搜到")
        try:
            # 尝试打开文件并发送
            return send_file(path, as_attachment=True)
        except PermissionError:
            return render_template("form/form_ERROR.html", id=session.get('userid'))
        except Exception as e:
            return render_template("form/form_ERROR.html", id=session.get('userid'))


# 开始运行
if __name__ == '__main__':
    # app.run()
    server = pywsgi.WSGIServer(('0.0.0.0', 6006), app)
    server.serve_forever()
